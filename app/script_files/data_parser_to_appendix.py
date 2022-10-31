#pip install openpyxl
#python3 script_files/data_parser_to_appendix.py uscc 1 2022  ---> from WSL
#-------------------
import sys
import pandas as pd
import sqlite3
pd.set_option('display.max_rows', 20)
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import time

start_time = time.time()


########Vars##############
cust_id = sys.argv[1]
month = int(sys.argv[2])
year = int(sys.argv[3])
path_d = f'./database_files/{cust_id}/'
path_r = f'./report_files/{cust_id}/'
##########################


con = sqlite3.connect(path_d + 'database_'+cust_id+'.sqlite')
#data = pd.read_csv("C:\\DATA\\Scripts\\Pull data from Vision by Marcelo\\AutoReport_v3.9\\report_files\\USCC\\USCC 10 2021 attacks.csv", parse_dates=['startDate','endDate'], dtype={"name": "string","attackIpsId":"string","actionType":"string","risk":"string"})
data = pd.read_sql_query("SELECT * from attacks", con)
con.close()

data_month = data[(data.month == month)] # data for the speicific month

data_month.to_csv(path_r+'database_'+cust_id+'_'+str(month)+'_'+str(year)+'.csv', encoding='utf-8', index=False)

## Events per Device (pie chart)##########################

s_epd = data_month.groupby(['deviceName','ruleName','name']).size()
s_epd  = s_epd.groupby(level=['deviceName'],group_keys=False).nlargest(5).to_frame('Top N Events Count')#.sort_values(by=['Top N Events Count'], ascending=False)
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx') as writer:
    s_epd .to_excel(writer, sheet_name='Events per Device')
print('Events per Device - complete')

## Malicious packets per Device##########################
s_mppd = data_month.groupby(['deviceName','ruleName','name']).sum()['packetCount']
s_mppd = s_mppd.groupby(level=['deviceName'],group_keys=False).nlargest(5).to_frame('Top N Malicious Packets')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_mppd.to_excel(writer, sheet_name='Malicious Packets per Device')
print('Malicious packets per Device- complete')

### Malicious bandwidth per Device#######################
pd.options.display.float_format = "{:,.2f}".format
s_mbpd = data_month.groupby(['deviceName','ruleName','name']).sum()['packetBandwidth']/8000000
s_mbpd = s_mbpd.groupby(level=['deviceName'],group_keys=False).nlargest(5).to_frame('Top N Malicious Bandwidth')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_mbpd.to_excel(writer, sheet_name='Malicious Bandwidth per Device')
print('Malicious bandwidth per Device - complete')

## Events per Policy######################################

s_epp = data_month.groupby(['ruleName','deviceName','name']).size()
s_epp  = s_epp .groupby(level=['ruleName'],group_keys=False).nlargest(5).to_frame('Top N Events Count')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_epp .to_excel(writer, sheet_name='Events per Policy')
print('Events per Policy - complete')

## Malicious packets per Policy############################
s_mppp = data_month.groupby(['ruleName','deviceName','name']).sum()['packetCount']
s_mppp = s_mppp.groupby(level=['ruleName'],group_keys=False).nlargest(5).to_frame('Top N Malicious Packets')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_mppp.to_excel(writer, sheet_name='Malicious Packets per Policy')
print('Malicious packets per Policy - complete')

## Malicious bandwidth per Policy#######################
pd.options.display.float_format = "{:,.2f}".format
s_mbpp = data_month.groupby(['ruleName','deviceName','name']).sum()['packetBandwidth']/8000000
s_mbpp = s_mbpp.groupby(level=['ruleName'],group_keys=False).nlargest(5).to_frame('Top N Malicious Bandwidth')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_mbpp.to_excel(writer, sheet_name='Malicious Bandwidth per Policy')
print('Malicious bandwidth per Policy - complete')

#Events per Destination IP Address######################

s_edip = data_month.groupby(['destAddress','name']).size()
s_edip  = s_edip.groupby(level=['destAddress'],group_keys=False).nlargest(5).to_frame('Top N Events Count')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_edip .to_excel(writer, sheet_name='Events per Dst IP')
print('Events per Destination IP Address - complete')

## Events per Attack Name ##############################

s_epan = data_month.groupby(['name','deviceName','ruleName']).size()
s_epan  = s_epan.groupby(level=['name'],group_keys=False).nlargest(5).to_frame('Top N Events Count')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_epan .to_excel(writer, sheet_name='Events per Attack Name')
print('Events per Attack Name - complete')

## Malicious packets per Attack Name####################

s_ppan = data_month.groupby(['name','deviceName','ruleName']).sum()['packetCount']
s_ppan = s_ppan.groupby(level=['name'],group_keys=False).nlargest(5).to_frame('Top N Malicious Packets')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_ppan.to_excel(writer, sheet_name='Malicious Pkts per Attack Name')
print('Malicious packets per Attack Name - complete')

## Malicious bandwidth per Attack Name##################

pd.options.display.float_format = "{:,.2f}".format
s_bpan = data_month.groupby(['name','deviceName','ruleName']).sum()['packetBandwidth']/8000000
s_bpan = s_bpan.groupby(level=['name'],group_keys=False).nlargest(5).to_frame('Top N Malicious Bandwidth')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_bpan.to_excel(writer, sheet_name='Malicious BW per Attack Name')
print('Malicious bandwidth per Attack Name - complete')

## Events per Destination Port##########################

s_epdp = data_month.groupby(['destPort','protocol','deviceName','ruleName','name']).size()
s_epdp  = s_epdp.groupby(level=['destPort'],group_keys=False).nlargest(10).to_frame('Top N Events Count')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_epdp .to_excel(writer, sheet_name='Events per Dst Port')
print('Events per Destination Port - complete')

## Events per protocol##################################

s_epp = data_month.groupby(['protocol','destPort','name']).size()
s_epp  = s_epp.groupby(level=['protocol'],group_keys=False).nlargest(10).to_frame('Top N Events Count')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_epp .to_excel(writer, sheet_name='Events per Protocol')
print('Events per protocol - complete')

## Malicious packets per Protocol#######################

s_ppp = data_month.groupby(['protocol','name']).sum()['packetCount']
s_ppp = s_ppp.groupby(level=['protocol'],group_keys=False).nlargest(5).to_frame('Top N Malicious Packets')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_ppp.to_excel(writer, sheet_name='Malicious Packets per Protocol')
print('Malicious packets per Protocol - complete')

## Malicious bandwidth per Protocol#####################

pd.options.display.float_format = "{:,.2f}".format
s_bpp = data_month.groupby(['protocol','name']).sum()['packetBandwidth']/8000000
s_bpp = s_bpp.groupby(level=['protocol'],group_keys=False).nlargest(5).to_frame('Top N Malicious Bandwidth')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_bpp.to_excel(writer, sheet_name='Malicious BW per Protocol')
print('Malicious bandwidth per Protocol - complete')

## Events per Source IP Address#########################

s_epsip = data_month.groupby(['sourceAddress','name']).size()
s_epsip  = s_epsip.groupby(level=['sourceAddress'],group_keys=False).nlargest(10).to_frame('Top N Events Count')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_epsip .to_excel(writer, sheet_name='Events per SrcIP')
print('Events per Source IP Address - complete')


## Packets per Source IP Address#########################

s_ppsip = data_month.groupby(['sourceAddress','name']).sum()['packetCount']
s_ppsip  = s_ppsip.groupby(level=['sourceAddress'],group_keys=False).nlargest(10).to_frame('Top N Packets Count')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_ppsip .to_excel(writer, sheet_name='Packets per SrcIP')
print('Packets per Source IP Address - complete')

## Bandwidth per Source IP Address#########################

s_bwpsip = data_month.groupby(['sourceAddress','name']).sum()['packetBandwidth']/8000000
s_bwpsip  = s_bwpsip.groupby(level=['sourceAddress'],group_keys=False).nlargest(10).to_frame('Top N BW')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_bwpsip .to_excel(writer, sheet_name='BW per SrcIP')
print('BW per Source IP Address - complete')
## Events per Source Country############################

s_epsc = data_month.groupby(['geoLocation','name']).size()
s_epsc  = s_epsc.groupby(level=['geoLocation'],group_keys=False).nlargest(10).to_frame('Top N Events Count')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_epsc .to_excel(writer, sheet_name='Events per Source Country')
print('Events per Source Country - complete')

## Events per Attack Duration###########################

s_epad = data_month.groupby(['durationRange','name']).size()
s_epad  = s_epad.groupby(level=['durationRange'],group_keys=False).nlargest(10).to_frame('Top N Events Count')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_epad.to_excel(writer, sheet_name='Events per Attack Duration')
print('Events per Attack Duration - complete')

## Events per Day of the Month##########################

s_epdm = data_month.groupby(['startDayOfMonth','name']).size()
s_epdm  = s_epdm.groupby(level=['startDayOfMonth'],group_keys=False).nlargest(10).to_frame('Top N Events Count')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
   s_epdm.to_excel(writer, sheet_name='Events per Day of Month')
print('vents per Day of the Month - complete')

## Malicious packets per Category#######################

s_ppc = data_month.groupby(['category','name']).sum()['packetCount']
s_ppc = s_ppc.groupby(level=['category'],group_keys=False).nlargest(5).to_frame('Top N Malicious Packets')#.to_csv('count_by_device-and_by_attack_name_full.csv', encoding='utf-8')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_ppc.to_excel(writer, sheet_name='Malicious Packets per Category')
print('Malicious packets per Category - complete')

## Malicious bandwidth per Category#####################

pd.options.display.float_format = "{:,.2f}".format
s_bpc = data_month.groupby(['category','name']).sum()['packetBandwidth']/8000000
s_bpc = s_bpc.groupby(level=['category'],group_keys=False).nlargest(5).to_frame('Top N Malicious Bandwidth')
with pd.ExcelWriter(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx', mode="a" ,if_sheet_exists ="replace") as writer:
    s_bpc.to_excel(writer, sheet_name='Malicious BW per Category')
print('Malicious bandwidth per Category - complete')

## Top #1 attack by packet count #######################


device = f"Device: {data_month.sort_values(by=['packetCount'], ascending=False).iloc[0]['deviceName']}"
policy = f"Policy: {data_month.sort_values(by=['packetCount'], ascending=False).iloc[0]['ruleName']}"
attack_type = f"Attack type: {data_month.sort_values(by=['packetCount'], ascending=False).iloc[0]['name']}"
src_ip = f"Source IP: {data_month.sort_values(by=['packetCount'], ascending=False).iloc[0]['sourceAddress']}"
dst_ip = f"Destination IP: {data_month.sort_values(by=['packetCount'], ascending=False).iloc[0]['destAddress']}"
attack_bw = f"Total attack traffic(cumulative in GB): {((data_month.sort_values(by=['packetCount'], ascending=False).iloc[0]['packetBandwidth'])/8000000)}"
attack_pkt = f"Total attack packets(cumulative): {format((data_month.sort_values(by=['packetCount'], ascending=False).iloc[0]['packetCount']), ',d')}"
duration = (data_month.sort_values(by=['packetCount'], ascending=False).iloc[0]['endTime'])-(data_month.sort_values(by=['packetCount'], ascending=False).iloc[0]['startTime'])
dur = f"Duration: {str(duration/60000) + ' min /' + str(duration/1000) + ' sec'}"
attack_start = f"Attack start time(UTC): {data_month.sort_values(by=['packetCount'], ascending=False).iloc[0]['startDate']}"
attack_id = f"Attack ID: {data_month.sort_values(by=['packetCount'], ascending=False).iloc[0]['attackIpsId']}"

#print(device)
#print(policy)
#print(attack_type)
#print(src_ip)
#print(dst_ip)
#print(attack_bw)
#print(attack_pkt)
#print(dur)
#print(attack_start)
#print(attack_id)

workbook_name = path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx'
wb = load_workbook(workbook_name)
wb.create_sheet("Top 1 attack by packet count")
page = wb["Top 1 attack by packet count"]

page.append([str(device)])
page.append([str(policy)])
page.append([str(attack_type)])
page.append([str(src_ip)])
page.append([str(dst_ip)])
page.append([str(attack_bw)])
page.append([str(attack_pkt)])
page.append([str(dur)])
page.append([str(attack_start)])
page.append([str(attack_id)])

wb.save(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx')

#Export the top attacks table by packet count to CSV
#topattackbypacketcount = data_month.sort_values(by=['packetCount'], ascending=False).iloc[0]['name']
#data_month[data_month['name'] == topattackbypacketcount].sort_values(by=['packetCount'], ascending=False).to_csv('top_attack_by_packetcount.csv', encoding='utf-8', index=False)
print('Top #1 attack by packet count - complete')

## Top #1 attack by bandwidth###########################

device = f"Device: {data_month.sort_values(by=['packetBandwidth'], ascending=False).iloc[0]['deviceName']}"
policy = f"Policy: {data_month.sort_values(by=['packetBandwidth'], ascending=False).iloc[0]['ruleName']}"
attack_type = f"Attack type: {data_month.sort_values(by=['packetBandwidth'], ascending=False).iloc[0]['name']}"
src_ip = f"Source IP: {data_month.sort_values(by=['packetBandwidth'], ascending=False).iloc[0]['sourceAddress']}"
dst_ip = f"Destination IP: {data_month.sort_values(by=['packetBandwidth'], ascending=False).iloc[0]['destAddress']}"
attack_bw = f"Total attack traffic(cumulative in GB): {((data_month.sort_values(by=['packetBandwidth'], ascending=False).iloc[0]['packetBandwidth'])/8000000)}"
attack_pkt = f"Total attack packets(cumulative): {format((data_month.sort_values(by=['packetBandwidth'], ascending=False).iloc[0]['packetCount']), ',d')}"
duration = (data_month.sort_values(by=['packetBandwidth'], ascending=False).iloc[0]['endTime'])-(data_month.sort_values(by=['packetBandwidth'], ascending=False).iloc[0]['startTime'])
dur = f"Duration: {str(duration/60000) + ' min /' + str(duration/1000) + ' sec'}"
attack_start = f"Attack start time(UT): {data_month.sort_values(by=['packetBandwidth'], ascending=False).iloc[0]['startDate']}"
attack_id = f"Attack ID: {data_month.sort_values(by=['packetBandwidth'], ascending=False).iloc[0]['attackIpsId']}"

#print(policy)
#print(device)
#print(attack_type)
#print(src_ip)
#print(dst_ip)
#print(attack_bw)
#print(attack_pkt)
#print(dur)
#print(attack_start)
#print(attack_id)

workbook_name = path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx'
wb = load_workbook(workbook_name)
wb.create_sheet("Top 1 attack by bandwidth")
page = wb["Top 1 attack by bandwidth"]

page.append([str(device)])
page.append([str(policy)])
page.append([str(attack_type)])
page.append([str(src_ip)])
page.append([str(dst_ip)])
page.append([str(attack_bw)])
page.append([str(attack_pkt)])
page.append([str(dur)])
page.append([str(attack_start)])
page.append([str(attack_id)])
    
wb.save(path_r+'appendix_'+cust_id+'_'+str(month)+'_'+str(year)+'.xlsx')

#Export the top attacks table by bandwidth to CSV

#topattackbybandwidth = data_month.sort_values(by=['packetBandwidth'], ascending=False).iloc[0]['name']
#data_month[data_month['name'] == topattackbybandwidth].sort_values(by=['packetBandwidth'], ascending=False).to_csv('top_attack_by_bandwidth.csv', encoding='utf-8', index=False)
print('Top #1 attack by bandwidth - complete')

print("--- %s seconds ---" % (time.time() - start_time))
